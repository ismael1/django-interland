import json
from django.db.models.query_utils import Q
from django.shortcuts import render
from django.http import Http404, FileResponse

# //verficar esta parte ismael 260521
# from rest_framework.response import HttpResponse
from django.http import JsonResponse, HttpResponse
from django.core.serializers import serialize
from django.http import JsonResponse
from django.template.response import TemplateResponse

import requests
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.decorators import api_view

from .models import Customer, Country, Estates, ZipCodes, RutasCustomer, TipoEmpresa, UsoCfdi, FormaPago, MetodoPago, AreaContacto, Contacto,FilesCustomer, DataComplementary, Business, Lada, Select_Zonas, Zonas, ZonasHijos
from .serializers import FiltroCustomerSerializer, RutaSerializer,ZipCodesSerializer, ContactoSerializer,ContactosSerializer, CustomersSerializer, CustomerSerializer, CountrySerializer, EstatesSerializer, TipoEmpresaSerializer, UsoCfdiSerializer, FormaPagoSerializer,MetodoPagoSerializer, AreaContactoSerializer,FilesCustomerSerializer, FilesSerializer, DataComplementarySerializer, BusinessSerializer, LadaSerializer, RutasSerializer, SelectZonasSerializer, ZonasSerializer, ZonasHijosSerializer

from catalogos.models import Geocercas
from catalogos.serializers import GeocercasSerializer

from rest_framework import viewsets
from rest_framework.authentication import BasicAuthentication
from rest_framework.permissions import IsAuthenticated

from django.db.models import Count

# from django.views.decorators.clickjacking import xframe_options_exempt 
# @xframe_options_exempt

# from django.views.decorators.clickjacking import xframe_options_exempt
from django.views.decorators.clickjacking import xframe_options_sameorigin
from django.views.decorators.csrf import csrf_exempt
from django.core.paginator import Paginator

import shutil

from django.core.files.storage import FileSystemStorage

from django.views.generic import TemplateView

###
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from django.template.loader import render_to_string
from django.conf import settings


#Generar PDF Librerias Necesarias Agregado David
from django.template.loader import render_to_string
import os
from django.views.generic import ListView, View 
from django.conf import settings
from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
from django.contrib.staticfiles import finders
#Fin Librerias PDF Necesarias

#Generar Excel Librerias Necesarias Agregado David
from django.http.response import HttpResponse
from django.views.generic.base import TemplateView
from openpyxl import Workbook, workbook
from openpyxl.styles import Alignment,Border,Font,PatternFill,Side
#Fin Librerias Excel Necesarias


# @xframe_options_exempt
# def ok_to_load_in_a_frame(request):
    # return HttpResponse("This page is safe to load in a frame on any site.")

class CustomersViewSet(viewsets.ModelViewSet):
    authentication_classes = (BasicAuthentication,)
    permission_classes = (IsAuthenticated,)
    queryset = Customer.objects.all()
    serializer_class = CustomersSerializer

class ContactosViewSet(viewsets.ModelViewSet):
    authentication_classes = (BasicAuthentication,)
    permission_classes = (IsAuthenticated,)
    queryset = Contacto.objects.all()
    serializer_class = ContactosSerializer

# @xframe_options_exempt
# @xframe_options_sameorigin

class FilesViewSet(viewsets.ModelViewSet): #Agregado David 310521
    authentication_classes = (BasicAuthentication,)
    permission_classes = (IsAuthenticated,)
    queryset = FilesCustomer.objects.all()
    # print (str(queryset.query))
    serializer_class = FilesSerializer

class DataComplementaryViewSet(viewsets.ModelViewSet): #Agregado David 310521
    authentication_classes = (BasicAuthentication,)
    permission_classes = (IsAuthenticated,)
    queryset = DataComplementary.objects.all()
    serializer_class = DataComplementarySerializer


##agregad 020621
class RutaViewSet(viewsets.ModelViewSet):
    authentication_classes = (BasicAuthentication,)
    permission_classes = (IsAuthenticated,)
    queryset = RutasCustomer.objects.all()
    serializer_class = RutaSerializer

#agregad  070621
class PaisesViewSet(viewsets.ModelViewSet):
    authentication_classes = (BasicAuthentication,)
    permission_classes = (IsAuthenticated,)
    queryset = Country.objects.all()
    serializer_class = CountrySerializer

class EstadosViewSet(viewsets.ModelViewSet):
    authentication_classes = (BasicAuthentication,)
    permission_classes = (IsAuthenticated,)
    queryset = Estates.objects.all()
    serializer_class = EstatesSerializer

class ListRutas(APIView):
    # country_o
    # estate_o
    # country_d
    # estate_d
    # def get_object(self, idCliente):
    #     try:
    #         # resul=RutasCustomer.objects.select_related('country_o', 'estate_o', 'country_d', 'estate_d', customer=idCliente)
    #         resul=RutasCustomer.objects.select_related('paisOrigen', 'estadoOrigen', 'paisDestino', 'estadoDestino')
    #         red = str(resul.query)
    #         # return RutasCustomer.objects.filter(customer=idCliente)
    #         return red
    #     except RutasCustomer.DoesNotExist:
    #         return Http404
    
    # def get(self, request, idCliente, format=None):        
    #     ruta = self.get_object(idCliente)
    #     serializer = RutasSerializer(ruta,many=True)
    #     return Response(serializer.data)
    def get_object(self, idCliente):
        try:
            # resul=RutasCustomer.objects.select_related('country_o', 'estate_o', 'country_d', 'estate_d').filter(customer=idCliente)
            # resul=RutasCustomer.objects.select_related('paisOrigen', 'estadoOrigen', 'paisDestino', 'estadoDestino').filter(customer=idCliente)
            # resul=RutasCustomer.objects.select_related('paisOrigen_id', 'paisDestino_id')
            # red = str(resul.query)
            resul = RutasCustomer.objects.filter(customer=idCliente)
            # resul = RutasCustomer.objects.filter(customer=idCliente)
            # print (str(resul.query))
            return resul
        except RutasCustomer.DoesNotExist:
            return Http404
    
    def get(self, request, idCliente, format=None):        
        ruta = self.get_object(idCliente)
        serializer = RutasSerializer(ruta,many=True)
        return Response(serializer.data)


class ListDataComplementary(APIView): #Agregado David 310521
    def get_object(self, idCliente):
        # print(idCliente)
        try:
            return DataComplementary.objects.get(customer=idCliente)
        except DataComplementary.DoesNotExist:
            return Http404

    def get(self, request, idCliente, format=None):
        
        datacomplementary = self.get_object(idCliente)
        serializer = DataComplementarySerializer(datacomplementary)
        return Response(serializer.data)

class DataComplementaryDetail(APIView): #Agregado David 310521
    
    def get_object(self, idCliente):
        # print(idCliente)
        try:
            return DataComplementary.objects.get(customer=idCliente)
        except DataComplementary.DoesNotExist:
            return Http404
    
    def get(self, request, idCliente, format=None):
        
        datacomplementary = self.get_object(idCliente)
        serializer = DataComplementarySerializer(datacomplementary)
        return Response(serializer.data)

class ListCustomer(APIView):
    def get(self, request, format=None):
        customers = Customer.objects.all()
        serializer = CustomerSerializer(customers, many=True)
        return Response(serializer.data)



@csrf_exempt 
@api_view(['GET'])
def ListCustomerFiltro(request):
# class ListCustomerFiltro(APIView): #Agregado David 100621
    # def get(self, request, format=None):
    #     customers = Customer.objects.all()
    #     serializer = FiltroCustomerSerializer(customers, many=True)
    #     return Response(serializer.data)
    # def get(self,request):
        # Get all the data in the table
        # current page
        page = int(request.GET.get('page', 1))
        # How many products are on a page
        size = int(request.GET.get('size', 1))

        #nuevo
        palabra=str(request.GET.get('palabra', 1))


         # Define where to start slicing
        data_start = (page - 1) * size

        # Define where to cut
        data_end = page * size

        # print (palabra)
        if (palabra == ""): 
            #Query data
            goodslist=Customer.objects.all()[data_start:data_end]
            count=Customer.objects.count()

        else:   
            palabra_min=palabra.lower()
            palabra_may=palabra.upper()
            goodslist = Customer.objects.filter(Q(name__icontains=palabra_min) | Q(name__icontains=palabra_may))[data_start:data_end]
            count=Customer.objects.filter(Q(name__icontains=palabra_min) | Q(name__icontains=palabra_may))[data_start:data_end].count()

        # Serialization operation
        goods_ser=FiltroCustomerSerializer(goodslist,many=True)

        # Return response
        return Response({
            'total':count,
            'data':goods_ser.data
        })

        # customers = Customer.objects.all()
        
        # Instantiated paging data, a data per page
        # pages = Paginator(customers,1)
        # # Add page numbers to the list
        # page_list = list()
        # for i in pages.page_range:
        #     page_list.append(i)
        
        # # The index parameter passed
        # paged = pages.page(index)

        # # Serialize the subscript
        # img_ser = FiltroCustomerSerializer(customers,many=True)
        # return Response({
        #     'code':200,
        #     'sum_page':pages.num_pages,
        #     'page_list':page_list,
        #     'data':img_ser.data
        # })



class CustomerDetail(APIView):    
    
    def get_object(self, idCliente):
        # print(idCliente)
        try:
            return Customer.objects.get(id=idCliente)
        except Customer.DoesNotExist:
            return Http404
    
    def get(self, request, idCliente, format=None):
        
        customer = self.get_object(idCliente)
        serializer = CustomerSerializer(customer)
        return Response(serializer.data)

class ListBusiness(APIView): #Agregado David 310521
    def get(self, request, format=None):
        bisness = Business.objects.all()
        serializer = BusinessSerializer(bisness, many=True)
        return Response(serializer.data)

class ListLada(APIView): #Agregado David 310521
    def get(self, request, format=None):
        ladas = Lada.objects.all()
        serializer = LadaSerializer(ladas, many=True)
        return Response(serializer.data)

        

#listado de paises
class ListCountry(APIView):
    def get(self, request, format=None):
        country = Country.objects.all()
        serializer = CountrySerializer(country, many=True)
        return Response(serializer.data)

#listado de estados
class ListEstates(APIView):    
    
    def get_object(self, idPais):
        try:
            return Estates.objects.filter(country_id=idPais)
        except Estates.DoesNotExist:
            return Http404
    
    def get(self, request, idPais, format=None):        
        estate = self.get_object(idPais)
        serializer = EstatesSerializer(estate,many=True)
        return Response(serializer.data)


#listado de empresas
class ListEmpresas(APIView):
    def get(self, request, format=None):
        empresas = TipoEmpresa.objects.all()
        serializer = TipoEmpresaSerializer(empresas, many=True)
        return Response(serializer.data)


#listado de usocfdi
class ListUsoCfdi(APIView):
    def get(self, request, format=None):
        usocfdi = UsoCfdi.objects.all()
        serializer = UsoCfdiSerializer(usocfdi, many=True)
        return Response(serializer.data)

#listado de formapago
class ListFormaPago(APIView):
    def get(self, request, format=None):
        formapago = FormaPago.objects.all()
        serializer = FormaPagoSerializer(formapago, many=True)
        return Response(serializer.data)

#listado de metodopago
class ListMetodoPago(APIView):
    def get(self, request, format=None):
        metodopago = MetodoPago.objects.all()
        serializer = MetodoPagoSerializer(metodopago, many=True)
        return Response(serializer.data)

#listado de areas
class ListAreas(APIView):
    def get(self, request, format=None):
        area = AreaContacto.objects.all()
        serializer = AreaContactoSerializer(area, many=True)
        return Response(serializer.data)

##Para el listado de contactos Ismael 260521
class ListContacts(APIView):

    def get_object(self, idCliente):
        try:
            return Contacto.objects.filter(customer=idCliente)
        except Contacto.DoesNotExist:
            return Http404
    
    def get(self, request, idCliente, format=None):        
        contacto = self.get_object(idCliente)
        serializer = ContactoSerializer(contacto,many=True)
        return Response(serializer.data)

#para los detalles del contacto ismael 260521
class ContactDetail(APIView):

    def get_object(self, idContact):
        # print(idCliente)
        try:
            return Contacto.objects.get(id=idContact)
        except Contacto.DoesNotExist:
            return Http404
    
    def get(self, request, idContact, format=None):
        
        contacto = self.get_object(idContact)
        serializer = ContactoSerializer(contacto)
        return Response(serializer.data)


##guardar archivos Ismael 260521
@csrf_exempt 
@api_view(['POST', 'GET'])
def addFiles(request):
    

    files_ = request.FILES.get('files_')
    idCustomer = request.POST.get('idCustomer')
    tipo = request.POST.get('tipo')

    # nom=request.FILES.get('files_').tmp_name

    # print (nom)
    # ///para compiar el archivo
    # fuente = files_
    # destino = "C:/Users/LATIN/OneDrive/Escritorio/SystemDev/InlandVueSystem/public/uploads/prueba.pdf"
    # shutil.copyfile(fuente, destino) 
    # myfile =  request.FILES.get('files_')
    # fs = FileSystemStorage()
    # fs = FileSystemStorage()
    # fs="C:/Users/LATIN/OneDrive/Escritorio/SystemDev/InlandVueSystem/public/uploads/prueba.pdf""
    # filename = fs.save(myfile.name, myfile)
    # uploaded_file_url = fs.url(filename)

    goods_obj = FilesCustomer.objects.create(name=files_, ruta=files_, tipo=tipo, statuspdf=0, customer_id=idCustomer)
    serializer = FilesCustomerSerializer(goods_obj)
    return Response(serializer.data)

    # goods_obj.save()
    # data = jsonResult.json_result(message="Added successfully", result="success", data=[], form_data={})
    # return Response(goods_obj.save())
    #  return HttpResponse(data)


#istar archivos ismael 260521
class ListFiles(APIView):
    def get_object(self, idCliente):
        try:
            return FilesCustomer.objects.filter(customer_id=idCliente)
        except FilesCustomer.DoesNotExist:
            return Http404
    
    def get(self, request, idCliente, format=None):        
        files = self.get_object(idCliente)
        serializer = FilesCustomerSerializer(files,many=True)
        return Response(serializer.data)

#istar archivos ismael 260521
class DetailsFiles(APIView):

    def get_object(self, idFile):
        # print(idCliente)
        try:
            return FilesCustomer.objects.get(id=idFile)
        except FilesCustomer.DoesNotExist:
            return Http404
    
    def get(self, request, idFile, format=None):
        
        files = self.get_object(idFile)
        serializer = FilesCustomerSerializer(files)
        return Response(serializer.data)


###agregado 270521
@csrf_exempt 
@api_view(['POST', 'GET'])
def showpDF(request):

    idF = request.POST.get('idFile')
# class showpDF(APIView):

    # def get_object(self, idFile):
        #Get the applicant's resume
    # ru='media/uploads/Anticipo_TRC-24.pdf'
    resume = FilesCustomer.objects.get(id=idF)
    ru2='media/' + str(resume.ruta)
    namef= str(resume.name)
    ru='http://127.0.0.1:8000/' + ru2
    # ru= ru2
    
    # ru='http://localhost:8080/logo_in.png'

    try:
        # return FileResponse(open(ru, 'rb'), content_type='application/pdf; charset=UTF-8')
        return HttpResponse(ru, content_type='application/pdf')
        # return HttpResponse(ru)

    except FileNotFoundError:
        raise Http404()

    # with open(ru, 'rb') as pdf:
    #     # file = open(str(pdf), encoding="utf8")
    #     response = HttpResponse(pdf.read(), content_type='application/pdf', encoding="utf8")
    #     response['Content-Disposition'] = 'inline;filename='+namef
    #     return response
    # pdf.closed

    # resume = FilesCustomer.objects.get(id=idF)
    # resume = FilesCustomer.object.filter(id=idF).filter(estatus='1')[0]
    # print(resume)
    # ru='media/' + str(resume.ruta)
    # namef= str(resume.name)

    # fsock = open(ru, 'r')
    # fsock = open(ru, 'r')

    # response = HttpResponse(fsock, content_type='application/pdf')
    # response = HttpResponse(fsock, content_type='application/pdf')

    # response['Content-Disposition'] = 'inline;filename='+namef
    # return response
    # return response
    
@csrf_exempt 
@api_view(['POST', 'GET'])
def searchZIP(request):
    dato = request.data.get('data')
    pais = request.data.get('pais')

#    .values('municipio')
    if dato:
        # zipCode = ZipCodes.objects.filter(codigo_postal=query)
        # zipCode = ZipCodes.objects.filter(codigo_postal__icontains=query).annotate(total=Count('codigo_postal'))
        zipCode = ZipCodes.objects.values('codigo_postal','municipio','estado_id','asentamiento').filter(codigo_postal__icontains=dato).filter(pais_id=pais).annotate(total=Count('codigo_postal'))



        # zipCode=ZipCodes.objects.values('id','codigo_postal','asentamiento','colonia','municipio','estado','pais').filter(codigo_postal__icontains=query).annotate(total=Count('municipio'))
        # zipCode=ZipCodes.objects.values('codigo_postal','municipio','estado_id','pais_id').filter(codigo_postal__icontains=query).annotate(total=Count('municipio','codigo_postal','estado_id','pais_id'))
        # zipCode=ZipCodes.objects.values('codigo_postal','municipio','estado_id','pais_id').filter(codigo_postal__icontains=query).annotate(total=Count('municipio','codigo_postal','estado_id','pais_id'))
        # zipCode=ZipCodes.objects.values('codigo_postal','municipio','estado','pais','id').filter(codigo_postal__icontains=query).annotate(total=Count('codigo_postal'))
        # zipCode=ZipCodes.objects.values('id','codigo_postal','municipio','estado','pais').filter(codigo_postal__icontains=query).annotate(total=Count('codigo_postal'))
        
        # red = str(zipCode.query)
        # print (red)
        # return Response({"no result": []})
        serializer = ZipCodesSerializer(zipCode, many=True)
        return Response(serializer.data)
    else:
        return Response({"no result": []})

@csrf_exempt 
@api_view(['POST', 'GET'])
def MunicipioViewSet(request):
    estado = request.data.get('estado')
    pais =   request.data.get('pais')
    
    if estado != '':
        
        cp = ZipCodes.objects.filter(pais_id = pais).filter(estado_id = estado).order_by("municipio").distinct("municipio")

        cpSer = ZipCodesSerializer(cp, many=True)
        return Response(cpSer.data)
    else:
        return Response({"no result": []})

@csrf_exempt 
@api_view(['POST', 'GET'])
def CodigoPostalViewSet(request):

    estado = request.data.get('estado')
    pais =   request.data.get('pais')
    muni =   request.data.get('municipio')

    ids = []
    listMun = []

    for mun in muni:
        ids.append(mun)

    municipio = ZipCodes.objects.filter(pais_id = pais).filter(estado_id = estado).filter(id__in = ids).order_by("municipio").distinct("municipio")

    for muni in municipio:
        listMun.append(muni.municipio)

    municipio = ZipCodes.objects.filter(pais_id = pais).filter(estado_id = estado).filter(municipio__in = listMun).order_by("municipio")
    munici = ZipCodesSerializer(municipio, many=True)
    
    return Response(munici.data)

@csrf_exempt 
@api_view(['POST', 'GET'])
def CodigoPostalCat(request):

    estado = request.data.get('estado')
    pais =   request.data.get('pais')
    muni =   request.data.get('municipio')

    municipio = ZipCodes.objects.filter(pais_id = pais).filter(estado_id = estado).filter(municipio = muni).order_by("asentamiento")
    munici = ZipCodesSerializer(municipio, many=True)
    
    return Response(munici.data) 

# Members.objects.values('designation').annotate(dcount=Count('designation'))

# SELECT designation, COUNT(designation) AS dcount
# FROM members GROUP BY designation
# y la salida sería de la forma

# [{'designation': 'Salesman', 'dcount': 2}, 
#  {'designation': 'Manager', 'dcount': 2}]

# @api_view(['POST'])
# def searchCustomer(request):
#     query = request.data.get('query', '')

#     if query:
#         custom = Customer.objects.filter(Q(name__icontains=query) | Q(estatus__icontains=query) | Q(dateCreate__icontains=query))
#         serializer = CustomerSerializer(custom, many=True)
#         return Response(serializer.data)
#     else:
#         return Response({"No hay datos": []})

# class TestView(TemplateView):

@csrf_exempt 
@api_view(['POST', 'GET'])
def send_email(request):

    pkC = request.data.get('id')


    try:
        # Establecemos conexion con el servidor smtp de gmail
        # mailServer = smtplib.SMTP(settings.EMAIL_HOST, settings.EMAIL_PORT)

        mensaje = MIMEMultipart()

        mailServer = smtplib.SMTP('smtp.office365.com', 587)
        print(mailServer.ehlo())
        mailServer.starttls()
        print(mailServer.ehlo())
        mailServer.login('infor@mx-interland.com', '=wAQ15&ix8')
        print('Conectado...')
        # mailServer = smtplib.SMTP(settings.EMAIL_HOST, settings.EMAIL_PORT)
        # print(mailServer.ehlo())
        # mailServer.starttls()
        # print(mailServer.ehlo())
        # mailServer.login(settings.EMAIL_HOST_USER, settings.EMAIL_HOST_PASSWORD)
        # print('Conectado...')

        #construye el mensaje simple
        email_to='danielrg841@gmail.com'

        
        # mensaje = MIMEText("""Este es el mensaje""")
        # mensaje['From']= settings.EMAIL_HOST_USER
        mensaje['From']= 'infor@mx-interland.com'
        mensaje['To']=email_to
        mensaje['Subject'] = "Solicitud de Credito"

        # render_to_string('send_email.html', {'solicitud': DataComplementary.objects.get(id=1)})
        # content = render_to_string('send_email.html', {'DataComplementary': DataComplementary.objects.filter(id=pkC)})
        # content = render_to_string('../templates/send_email.html')

        #  DataComplementary.objects.get(customer=idCliente)
        # mensaje.attach(MIMEText(content, 'html'))
        # content = render_to_string('send_email.html')
        # print(content)

        # print(pkC)

        content = render_to_string('send_email_credito.html', {'DataComplementary': DataComplementary.objects.get(id=pkC)})
        mensaje.add_header('Content-Type', 'text/html')
        mensaje.set_payload(content)

        # print(content)
        # mensaje.attach(MIMEText(content, 'text/html'))
        # mensaje.attach(MIMEText(content, 'text/html'))
        # mensaje.attach(content, 'text/html')        
        # part2 = MIMEText(content, 'text/html')

        # mensaje.attach_alternative(content, "text/html")        
        # print(mensaje)
        # mensaje.attach(MIMEText(content, 'html'))
        # mensaje.attach(MIMEText(content, 'html'))
        # print(mensaje)


        mailServer.sendmail('infor@mx-interland.com',
                             email_to,
                             mensaje.as_string())

        return Response({"msg": 'exito'})


    except Exception  as e:
        # print(e)
        return Response({"msg": e})


    # template_name ='send_email.html'

    # # @method_decorator(csrf_exempt)

    # def disparch(self, request, *args, **kwargs):
    #     return super().dispatch(request, *args, **kwargs)

    # # def post(self, request, *args, **kwargs):
    
    # def get_context_data(self, **kwargs):
    #     context = super().get_context_data(**kwargs)
    #     context['title'] = 'Select Anidados | Django'
    #     return context


#GENERAR PDF Y ECEL
class PDFCustomer(View):
    def get(self, request, *args, **kwargs):
        customers = Customer.objects.all()
        serializer = CustomerSerializer(customers, many=True)
        # return Response(serializer.data)
        # template = get_template('pdf/customers.html')
        template = render_to_string('pdf/pdf_customers.html', {'Customer': serializer.data})  
        # print(template)
        # context = {'title': 'Lista Clientes PDF'}
        # html = template.render(context)
        html = template
        response = HttpResponse(content_type= 'application/pdf')
        response['Content-Disposition'] = 'attachment; filename="Reporte Clientes.pdf"'
        pisa_status = pisa.CreatePDF(html, dest = response)
        if pisa_status.err:
            return HttpResponse('We had some errors <pre>' + html + '</pre>')
        return response

def ExcelCustomer(request):
    queryset = Customer.objects.all()
    return render(request, 'excel/excel_customers.html',{'queryset':queryset})

class ReportePersonalizadoExcel(View):
    def get(self,request,*args,**kwargs):
        query = Customer.objects.all()
        wb = Workbook()
        ws = wb.active
        ws.title = 'Lista Clientes'

        ws['B1'].alignment = Alignment(horizontal = "center", vertical = "center") #Asigna la Alineacion de la Celda en este caso Centrado
        ws['B1'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), 
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin")) #Crear Borde Izquierdo, Derecho, Superior, Inferior de la Celda B1
        ws['B1'].fill = PatternFill(start_color = 'C6C6C6', end_color = 'C6C6C6', fill_type = "solid") #Asigna color de fondo a la Celda B1
        ws['B1'].font = Font(name = 'Calibri', size = 12, bold = True) #Asigna Tipo de Fuente y Tamaño a la Celda B1
        ws['B1'] = 'Reporte de Clientes' #Escribe Texto al Excel

        ws.merge_cells('B1:E1') #Combina Celdas en este caso desde B1 asta E1
        ws.row_dimensions[1].height = 20 #Asigna la Altura de la Celda
        ws.column_dimensions['B'].width = 20 #Asigna el Ancho de la Columna B
        ws.column_dimensions['C'].width = 20 #Asigna el Ancho de la Columna C
        ws.column_dimensions['D'].width = 20 #Asigna el Ancho de la Columna D
        ws.column_dimensions['E'].width = 20 #Asigna el Ancho de la Columna E

        controlador = 4
        for q in query:  

            ws['B3'].alignment = Alignment(horizontal = "center", vertical = "center")
            ws['B3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
            ws['B3'].fill = PatternFill(start_color = 'C6C6C6', end_color = 'C6C6C6', fill_type = "solid")
            ws['B3'].font = Font(name = 'Calibri', size = 10, bold = True)
            ws['B3'] = 'Nombre'

            ws['C3'].alignment = Alignment(horizontal = "center", vertical = "center")
            ws['C3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
            ws['C3'].fill = PatternFill(start_color = 'C6C6C6', end_color = 'C6C6C6', fill_type = "solid")
            ws['C3'].font = Font(name = 'Calibri', size = 10, bold = True)
            ws['C3'] = 'Email'

            ws['D3'].alignment = Alignment(horizontal = "center", vertical = "center")
            ws['D3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
            ws['D3'].fill = PatternFill(start_color = 'C6C6C6', end_color = 'C6C6C6', fill_type = "solid")
            ws['D3'].font = Font(name = 'Calibri', size = 10, bold = True)
            ws['D3'] = 'Telefono'

            ws['E3'].alignment = Alignment(horizontal = "center", vertical = "center")
            ws['E3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
            ws['E3'].fill = PatternFill(start_color = 'C6C6C6', end_color = 'C6C6C6', fill_type = "solid")
            ws['E3'].font = Font(name = 'Calibri', size = 10, bold = True)
            ws['E3'] = 'Pais'

            #Llenar el Reporte
            ws.cell(row = controlador, column = 2).alignment = Alignment(horizontal = "center")
            ws.cell(row = controlador, column = 2).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
            ws.cell(row = controlador, column = 2).font = Font(name = 'Calibri', size = 8)
            ws.cell(row = controlador, column = 2).value = q.name

            ws.cell(row = controlador, column = 3).alignment = Alignment(horizontal = "center")
            ws.cell(row = controlador, column = 3).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
            ws.cell(row = controlador, column = 3).font = Font(name = 'Calibri', size = 8)
            ws.cell(row = controlador, column = 3).value = q.email

            ws.cell(row = controlador, column = 4).alignment = Alignment(horizontal = "center")
            ws.cell(row = controlador, column = 4).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
            ws.cell(row = controlador, column = 4).font = Font(name = 'Calibri', size = 8)
            ws.cell(row = controlador, column = 4).value = q.phone

            ws.cell(row = controlador, column = 5).alignment = Alignment(horizontal = "center")
            ws.cell(row = controlador, column = 5).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
            ws.cell(row = controlador, column = 5).font = Font(name = 'Calibri', size = 8)
            ws.cell(row = controlador, column = 5).value = q.pais

            controlador += 1
        
        #Establecer el nombre de archivo
        nombre_archivo = "Reporte Clientes.xlsx"
        #Definir el tipo de respuesta que se va a dar
        response = HttpResponse(content_type = "application/ms-excel")
        contenido = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response

#Agregado Lista Archivos
@csrf_exempt 
@api_view(['POST'])
def consultarFiles(request):
    
    customer = request.data.get('customer_id')
    tipofile = request.data.get('tipo')

    if customer:
        buscar = FilesCustomer.objects.filter(customer_id=customer).filter(tipo=tipofile)
        serializer = FilesCustomerSerializer(buscar, many=True)
        return Response(serializer.data)
    else:
        return Response({"no result": []})

#INICIA SECCION DE ZONAS

@csrf_exempt
@api_view(['GET'])
def ListZonasFiltro(request):

    page = int(request.GET.get('page', 1))
    size = int(request.GET.get('size', 1))
    palabra=str(request.GET.get('palabra', 1))
    data_start = (page - 1) * size
    data_end = page * size

    if (palabra == ""):
        goodslist=Zonas.objects.filter(Q(estatus=1)).order_by('llave').distinct('llave')[data_start:data_end]
        count=Zonas.objects.filter(Q(estatus=1)).order_by('llave').distinct('llave').count()
    else:
        goodslist = Zonas.objects.filter(Q(identificador__icontains=palabra) | Q(clasificacion__icontains=palabra)).filter(Q(estatus=1)).order_by('llave').distinct('llave')[data_start:data_end]
        count = goodslist.count()

    goods_ser=ZonasSerializer(goodslist,many=True)

    return Response({
        'total':count,
        'data':goods_ser.data
    })

@csrf_exempt 
@api_view(['POST', 'GET'])
def nuevaZona(request):

    id = request.data.get('id')
    clasificacion = request.data.get('clasificacion')
    claveidentificador = request.data.get('claveidentificador')
    estado = request.data.get('estado')
    identificador = request.data.get('identificador')
    municipio = request.data.get('municipio')
    nomEstado = request.data.get('nomEstado')
    pais = request.data.get('pais')

    llave = identificador + " " + claveidentificador

    idMun = []
    listZonas = []

    for mun in municipio:
        idMun.append(mun)

    municipio = ZipCodes.objects.filter(pais_id = pais).filter(estado_id = estado).filter(id__in = idMun).order_by("municipio").distinct("municipio")

    for nomMun in municipio:

        consulta = Zonas.objects.all().filter(clasificacion=clasificacion).filter(claveIde=claveidentificador).filter(identificador=identificador).filter(estado_n=nomEstado).filter(municipio=nomMun.municipio).filter(estado_id=estado).filter(pais_id=pais).filter(estatus=1)

        if(len(consulta) == 0 ):
            consecutivo = Zonas.objects.create(clasificacion=clasificacion, claveIde=claveidentificador, identificador=identificador, llave=llave ,estado_n=nomEstado, municipio=nomMun.municipio, estado_id=estado, pais_id=pais, estatus=1)
            serializer = ZonasSerializer(consecutivo)
            listZonas.append(serializer.data)

    return Response(listZonas)

@csrf_exempt 
@api_view(['POST'])
def getId(request):

    id = request.data.get('id')
    if id != 0:
        consecutivo = Zonas.objects.all().order_by("-idZona")[:1]

        serializer = ZonasSerializer(consecutivo, many=True)
        return Response(serializer.data)
    else:
        return Response({"no result": []})


@csrf_exempt
@api_view(['GET'])
def consultaHijos(request):

    page = int(request.GET.get('page', 1))
    size = int(request.GET.get('size', 1))
    palabra=str(request.GET.get('palabra', 1))
    data_start = (page - 1) * size
    data_end = page * size
    llave = request.GET.get('llave')
    
    print(page)
    print(size)
    print(palabra)
    print(data_start)
    print(data_end)
    print(llave)

    idZ = []

    ids = Zonas.objects.filter(Q(llave=llave))

    for z in ids:
        idZ.append(z.idZona)

    if (palabra == ""):

      
        goodslist=ZonasHijos.objects.filter(idZonas_id__in=idZ)[data_start:data_end]
        count=ZonasHijos.objects.filter(idZonas_id__in=idZ).count()
    else:
        goodslist = ZonasHijos.objects.filter(Q(colonia__icontains=palabra) | Q(municipio__icontains=palabra) | Q(codigoPostal__icontains=palabra)).filter(Q(idZonas_id__in=idZ))[data_start:data_end]
        count=goodslist.count()

    goods_ser=ZonasHijosSerializer(goodslist,many=True)

    return Response({
        'total':count,
        'data':goods_ser.data
    })

@csrf_exempt 
@api_view(['POST', 'GET'])
def nuevaZonaHijos(request):

    resp = request.data.get('responseZonas')
    cps = request.data.get('cps')

    listCP = []

    for i in resp['data']:
        for j in cps:
            valida = ZipCodes.objects.filter(id=j)
            val = ZipCodesSerializer(valida, many=True)
            if val.data[0]['municipio'] == i['municipio']:
                insert = ZonasHijos.objects.create(codigoPostal=val.data[0]['codigo_postal'], colonia=val.data[0]['asentamiento'], idZonas_id=i['idZona'], municipio=i['municipio'])
                ser = ZonasHijosSerializer(insert, many=True)
                listCP.append(insert.idZonasHijo)
                print(listCP)
    
    return Response(listCP)

@csrf_exempt
@api_view(['GET'])
def getZonas(request, idZona):

    zona = Zonas.objects.filter(Q(idZona=idZona)).order_by('idZona')
    zonaSer = ZonasSerializer(zona,many=True)

    hijos = ZonasHijos.objects.filter(Q(idZonas_id=idZona)).order_by('idZonasHijo')
    hijosSer = ZonasHijosSerializer(hijos,many=True)
        
    return Response({'zona':zonaSer.data, 'zonas': hijosSer.data})

@csrf_exempt 
@api_view(['PUT'])
def deleteZona(request):
    
    idS = request.data.get('idZona')

    if idS != 0:

        hijo = ZonasHijos.objects.filter(idZonas_id=idS).delete()

        Zona = Zonas.objects.filter(idZona=idS).delete()

        return Response(True)
    else:
        return Response(False)

@csrf_exempt 
@api_view(['PUT'])
def deleteHijo(request):
    
    idS = request.data.get('idZona')

    if idS != 0:

        hijo = ZonasHijos.objects.filter(idZonasHijo=idS).delete()

        return Response(True)
    else:
        return Response(False)


@csrf_exempt 
@api_view(['POST', 'GET'])
def getCoincidenciaZona(request):

    cpO = request.data.get('codPosO')
    cpD = request.data.get('codPosD')

    respO = []
    respD = []

    hijosO = ZonasHijos.objects.filter(Q(codigoPostal=cpO)).order_by('codigoPostal')
    hijosOSer = ZonasHijosSerializer(hijosO,many=True)

    if len(hijosO) > 0:
        zonaO = Zonas.objects.filter(Q(idZona=hijosOSer.data[0]['idZonas'])).order_by('idZona')
        zonaOSer = ZonasSerializer(zonaO,many=True)
        
        respO = zonaOSer.data

    hijosD = ZonasHijos.objects.filter(Q(codigoPostal=cpD)).order_by('codigoPostal')
    hijosDSer = ZonasHijosSerializer(hijosD,many=True)

    if len(hijosD) > 0:
        zonaD = Zonas.objects.filter(Q(idZona=hijosDSer.data[0]['idZonas'])).order_by('idZona')
        zonaDSer = ZonasSerializer(zonaD,many=True)
        
        respD = zonaDSer.data

    return Response({'origen': respO, 'destino':respD})

@csrf_exempt 
@api_view(['POST', 'GET'])
def validaCPGeocerca(request):

    cpO = request.data.get('codPosO')
    cpD = request.data.get('codPosD')

    respO = []
    respD = []
    estatusO = ""
    estatusD = ""

    zonaOrigen = Geocercas.objects.filter(estatus=1, codigoPostal = cpO).values('codigoPostal', 'estatus_geocerca').distinct()
    zonaOSer = GeocercasSerializer(zonaOrigen,many=True)

    if len(zonaOrigen) > 0:

        if zonaOSer.data[0]['estatus_geocerca'] == 1:
            estatusO = "COMERCIAL"
        elif zonaOSer.data[0]['estatus_geocerca'] == 2:
            estatusO = "NO COMERCIAL"
        elif zonaOSer.data[0]['estatus_geocerca'] == 3:
            estatusO = "PELIGROSA"
        elif zonaOSer.data[0]['estatus_geocerca'] == 4:
            estatusO = "RESTRINGIDA"

        respO = [{"cp":zonaOSer.data[0]['codigoPostal'], "estatus":estatusO}]

    zonaDestino = Geocercas.objects.filter(estatus=1, codigoPostal = cpD).values('codigoPostal', 'estatus_geocerca').distinct()
    zonaDSer = GeocercasSerializer(zonaDestino,many=True)

    if len(zonaDestino) > 0:

        if zonaDSer.data[0]['estatus_geocerca'] == 1:
            estatusD = "COMERCIAL"
        elif zonaDSer.data[0]['estatus_geocerca'] == 2:
            estatusD = "NO COMERCIAL"
        elif zonaDSer.data[0]['estatus_geocerca'] == 3:
            estatusD = "PELIGROSA"
        elif zonaDSer.data[0]['estatus_geocerca'] == 4:
            estatusD = "RESTRINGIDA"

        respD = [{"cp":zonaDSer.data[0]['codigoPostal'], "estatus":estatusD}]

    return Response({'origen': respO, 'destino':respD})

@csrf_exempt 
@api_view(['POST', 'GET'])
def coincidenciaZona(request):

    cp = request.data.get('cp')

    respO = False

    hijosO = ZonasHijos.objects.filter(Q(codigoPostal=cp)).order_by('codigoPostal')
    hijosOSer = ZonasHijosSerializer(hijosO,many=True)

    if len(hijosO) > 0:
        zonaO = Zonas.objects.filter(Q(idZona=hijosOSer.data[0]['idZonas'])).order_by('idZona')
        zonaOSer = ZonasSerializer(zonaO,many=True)
        
        respO = zonaOSer.data

    return Response(respO)

#TERMINA SECCION DE ZONAS

@csrf_exempt 
@api_view(['POST', 'GET'])
def searchAddress(request):
    dato = request.data.get('data')

    dato = dato.upper()

    datosUbi = []

    
    #ubicacion = ZipCodes.objects.filter(Q(direccion__icontains=dato)).select_related('pais','estado').values('pais__id','pais__code','pais__name','pais__estatus','estado__id','estado__code','estado__name','codigo_postal','asentamiento','municipio')[:15]
    ubicacion = Geocercas.objects.filter(Q(direccion__icontains=dato) | Q(nombre_corto__icontains=dato))[:25]
    ubi = GeocercasSerializer(ubicacion,many=True)
    #print(ubicacion.query)
   
    return Response({'data':ubi.data})


@csrf_exempt 
@api_view(['POST', 'GET'])
def searchAddressDetalles(request):
    dato = request.data.get('data')
    cp = request.data.get('cp')
    cd = request.data.get('cd')


    datosUbi = []

    #ubicacion = ZipCodes.objects.filter(Q(asentamiento__icontains=dato) | Q(codigo_postal__icontains=dato) | Q(municipio__icontains=dato) | Q(estado_n__icontains=dato)).select_related('pais','estado').values('pais__id','pais__code','pais__name','pais__estatus','estado__id','estado__code','estado__name','codigo_postal','asentamiento','municipio')
    ubicacion = ZipCodes.objects.filter(Q(direccion__icontains=dato),Q(codigo_postal__icontains=cp),Q(asentamiento__icontains=cd)).select_related('pais','estado').values('pais__id','pais__code','pais__name','pais__estatus','estado__id','estado__code','estado__name','codigo_postal','asentamiento','municipio')[:15]
    print(ubicacion.query)
    for u in ubicacion:
        
        
        datosUbi.append({
            "pais_id":u['pais__id'],
            "pais_code":u['pais__code'],
            "pais_name":u['pais__name'],
            "pais_estatus":u['pais__estatus'],
            "estado_id":u['estado__id'],
            "estado_code":u['estado__code'],
            "estado_name":u['estado__name'],
            "codigo_postal":u['codigo_postal'],
            "asentamiento":u['asentamiento'],
            "municipio":u['municipio'],
        })
    return Response(datosUbi)

@csrf_exempt
@api_view(['GET'])
def getSelectZonas(request):

    zona = Select_Zonas.objects.order_by('id')
    zonaSer = SelectZonasSerializer(zona,many=True)
        
    return Response({'zona':zonaSer.data})

@csrf_exempt 
@api_view(['POST', 'GET'])
def updateSelectZonas(request):

    id = int(request.data.get('id'))
    tipo = str(request.data.get('tipo'))
    estatus = bool(request.data.get('estatus'))
    print(estatus, tipo, id)

    zn = Select_Zonas.objects.get(id=id)
    if tipo == 'c':
        zn.check_zona_c = estatus
    
    if tipo == 'nc':
        zn.check_zona_nc = estatus

    if tipo == 'p':
        zn.check_zona_p = estatus

    zn.save()

    zonas = Select_Zonas.objects.filter(Q(id=id)).order_by('id')
    zonasSer = SelectZonasSerializer(zonas,many=True)

    
    respO = zonasSer.data

    return Response(respO)

@csrf_exempt 
@api_view(['POST','GET'])
def datos_select(request):

    info_update = "jeje"
    update = ""
    zonas = Select_Zonas.objects.all()

    for dato in zonas:
        update += "UPDATE customer_select_zonas SET check_zona_c = "+str(dato.check_zona_c)+",  check_zona_nc = "+str(dato.check_zona_nc)+", check_zona_p = "+str(dato.check_zona_p)+" where id = "+str(dato.id)+";\n"

    return Response({"info": update})

'''@csrf_exempt 
@api_view(['POST','GET'])
def zona_geocercas(request):

    tipo_zona = int(request.data.get('tz'))
    zonas = []
    arrayDirecciones = []
    #API_KEY = 'AIzaSyADhOxfxQ9u-0_4FuHs8sVMHnyw0TnI11Y'
    API_KEY = 'AIzaSyDdEvJBlhZkgA1qRF7pZen1--yH6FB2AfA'
    gpais = ''
    gestado = ''
    gciudad = ''
    gcolonia = ''
    gcp = ''
    glat = ''
    glng = ''
    cont = 0
    data = []
    idZonaAct = 0

    acentos = 'áéíóúÁÉÍÓÚñÑüÜ'
    sin_acentos = 'aeiouAEIOUnNuU'
    trans = str.maketrans(acentos, sin_acentos)
    
    if tipo_zona == 1:
        zonas = Select_Zonas.objects.filter(Q(check_zona_c=1)).filter(Q(migrate_zona_c=0))
    elif tipo_zona == 2:
        zonas = Select_Zonas.objects.filter(Q(check_zona_nc=1)).filter(Q(migrate_zona_nc=0))
    elif tipo_zona == 3:
        zonas = Select_Zonas.objects.filter(Q(check_zona_p=1)).filter(Q(migrate_zona_p=0))

    for zona in zonas:
        idZonaAct = zona.id
        cp = ZipCodes.objects.filter(Q(direccion__startswith=zona.direccion)).order_by("municipio")
        for dir in cp:
            
            direccion_codificada_origen = requests.utils.quote(dir.direccion)
    
            response_origen = requests.get(f"https://maps.googleapis.com/maps/api/geocode/json?address={direccion_codificada_origen}&key={API_KEY}")
            # Verificar si la solicitud fue exitosa
            if response_origen.status_code == 200:
                data = response_origen.json()
                #print(data)
                # Verificar si se encontró una ubicación
                if data.get('results') and len(data['results']) > 0:
                    # Obtener la dirección completa
                    #direccion_completa = data['results'][0]['formatted_address']
                    for estado in data['results'][0]['address_components']:
                        
                        for tipo in estado['types']:
                            if tipo == 'country':
                                gpais = estado['long_name']
                                gpais = gpais.title().translate(trans)

                        for tipo in estado['types']:
                            if tipo == 'administrative_area_level_1':
                                if estado['long_name'] == 'Ciudad de México' or estado['long_name'] == 'Mexico City' or estado['long_name'] == 'CDMX':
                                    gestado = 'Ciudad de Mexico'
                                elif  estado['long_name'] == 'Estado de México' or estado['long_name'] == 'State of Mexico' or estado['long_name'] == 'México':
                                    gestado = 'Estado de Mexico'
                                else:
                                    gestado = estado['long_name']
                                    gestado = gestado.title().translate(trans)

                        for tipo in estado['types']:
                            if tipo == 'locality':
                                gciudad = estado['long_name']
                                gciudad =gciudad.title().translate(trans)
                        
                        for tipo in estado['types']:
                            if tipo == 'sublocality' or tipo == 'sublocality_level_1':
                                gcolonia = estado['long_name']
                                gcolonia = gcolonia.title().translate(trans)

                        for tipo in estado['types']:
                            if tipo == 'postal_code':
                                gcp = estado['long_name']

                        glat = data['results'][0]['geometry']['location']['lat']
                        glng = data['results'][0]['geometry']['location']['lng']
                    
                    direccionCompletaOrigen = gpais+', '+gestado+', '+gciudad+', '+gcolonia+', '+gcp

                    
                else:
                    print("No se pudo encontrar una dirección para la ubicación proporcionada.")
            else:
                print("Error al realizar la solicitud:", response_origen.status_code)
            
            print(direccionCompletaOrigen, glat, glng, gpais, gestado, gciudad, gcolonia, gcp) #Obtenemos Direccion completa Origen

            if len(gcp) <= 10:
                insert = Geocercas.objects.create(codigoPostal = gcp, lat= glat, lng = glng, orden = 1, estatus = 1, poligono = 1, usuarioAlta = 'admin', pais = gpais, ciudad = gciudad, colonia = gcolonia, estado = gestado, lat_centro = glat, lng_centro = glng, estatus_geocerca = tipo_zona, kilometros_redonda = 5)
            
                act = Select_Zonas.objects.get(id = idZonaAct)
                if tipo_zona == 1:
                    act.migrate_zona_c = 1
                elif tipo_zona == 2:
                    act.migrate_zona_nc = 1
                elif tipo_zona == 3:
                    act.migrate_zona_p = 1
                act.save()

    return Response({"info": insert})'''

'''@csrf_exempt 
@api_view(['POST','GET'])
def zona_geocercas(request):

    tipo_zona = int(request.data.get('tz'))
    zonas = []
    arrayDirecciones = []
    API_KEY = 'AIzaSyADhOxfxQ9u-0_4FuHs8sVMHnyw0TnI11Y'
    #API_KEY = 'AIzaSyDdEvJBlhZkgA1qRF7pZen1--yH6FB2AfA'
    gpais = ''
    gestado = ''
    gciudad = ''
    gcolonia = ''
    gcp = ''
    glat = ''
    glng = ''
    cont = 0
    data = []
    idZonaAct = 0

    acentos = 'áéíóúÁÉÍÓÚñÑüÜ'
    sin_acentos = 'aeiouAEIOUnNuU'
    trans = str.maketrans(acentos, sin_acentos)
    
    if tipo_zona == 1:
        zonas = Select_Zonas.objects.filter(Q(check_zona_c=1)).filter(Q(migrate_zona_c=0))
    elif tipo_zona == 2:
        zonas = Select_Zonas.objects.filter(Q(check_zona_nc=1)).filter(Q(migrate_zona_nc=0))
    elif tipo_zona == 3:
        zonas = Select_Zonas.objects.filter(Q(check_zona_p=1)).filter(Q(migrate_zona_p=0))

    for zona in zonas:
        idZonaAct = zona.id 
        
        #print(zona.direccion)
        geo = Geocercas.objects.filter(Q(direccion__icontains=zona.direccion))
        

        if geo.count() > 0:
    
            if tipo_zona == 1:
                geo.update(estatus = 1)
            elif tipo_zona == 2:
                geo.update(estatus = 1)
            elif tipo_zona == 3:
                geo.update(estatus = 1)


            act = Select_Zonas.objects.get(id = idZonaAct)
            if tipo_zona == 1:
                act.migrate_zona_c = 1
            elif tipo_zona == 2:
                act.migrate_zona_nc = 1
            elif tipo_zona == 3:
                act.migrate_zona_p = 1
            act.save()
        
        else:
            cp = ZipCodes.objects.filter(Q(direccion__startswith=zona.direccion)).order_by("municipio")
            for dir in cp:
                direccion_codificada_origen = requests.utils.quote(dir.direccion)
                response_origen = requests.get(f"https://maps.googleapis.com/maps/api/geocode/json?address={direccion_codificada_origen}&key={API_KEY}")
                # Verificar si la solicitud fue exitosa
                if response_origen.status_code == 200:
                    data = response_origen.json()
                    #print(data)
                    # Verificar si se encontró una ubicación
                    if data.get('results') and len(data['results']) > 0:
                        # Obtener la dirección completa
                        #direccion_completa = data['results'][0]['formatted_address']
                        for estado in data['results'][0]['address_components']:
                            for tipo in estado['types']:
                                if tipo == 'country':
                                    gpais = estado['long_name']
                                    gpais = gpais.title().translate(trans)

                        for tipo in estado['types']:
                            if tipo == 'administrative_area_level_1':
                                if estado['long_name'] == 'Ciudad de México' or estado['long_name'] == 'Mexico City' or estado['long_name'] == 'CDMX':
                                    gestado = 'Ciudad de Mexico'
                                elif  estado['long_name'] == 'Estado de México' or estado['long_name'] == 'State of Mexico' or estado['long_name'] == 'México':
                                    gestado = 'Estado de Mexico'
                                else:
                                    gestado = estado['long_name']
                                    gestado = gestado.title().translate(trans)

                        for tipo in estado['types']:
                            if tipo == 'locality':
                                gciudad = estado['long_name']
                                gciudad =gciudad.title().translate(trans)
                        
                        for tipo in estado['types']:
                            if tipo == 'sublocality' or tipo == 'sublocality_level_1':
                                gcolonia = estado['long_name']
                                gcolonia = gcolonia.title().translate(trans)

                        for tipo in estado['types']:
                            if tipo == 'postal_code':
                                gcp = estado['long_name']

                        glat = data['results'][0]['geometry']['location']['lat']
                        glng = data['results'][0]['geometry']['location']['lng']
                    
                    direccionCompletaOrigen = gpais+', '+gestado+', '+gciudad+', '+gcolonia+', '+gcp
                else:
                    print("No se pudo encontrar una dirección para la ubicación proporcionada.")
            else:
                print("Error al realizar la solicitud:", response_origen.status_code)
            
            print(direccionCompletaOrigen, glat, glng, gpais, gestado, gciudad, gcolonia, gcp) #Obtenemos Direccion completa Origen

            if len(gcp) <= 10:
                insert = Geocercas.objects.create(codigoPostal = gcp, lat= glat, lng = glng, orden = 1, estatus = 1, poligono = 1, usuarioAlta = 'admin', pais = gpais, ciudad = gciudad, colonia = gcolonia, estado = gestado, lat_centro = glat, lng_centro = glng, estatus_geocerca = tipo_zona, kilometros_redonda = 5)
            
                act = Select_Zonas.objects.get(id = idZonaAct)
                if tipo_zona == 1:
                    act.migrate_zona_c = 1
                elif tipo_zona == 2:
                    act.migrate_zona_nc = 1
                elif tipo_zona == 3:
                    act.migrate_zona_p = 1
                act.save()

    return Response({"info": ''})'''


@csrf_exempt 
@api_view(['POST','GET'])
def zona_geocercas(request):

    tipo_zona = int(request.data.get('tz'))
    zonas = []
    arrayDirecciones = []
    API_KEY = 'AIzaSyADhOxfxQ9u-0_4FuHs8sVMHnyw0TnI11Y'
    #API_KEY = 'AIzaSyDdEvJBlhZkgA1qRF7pZen1--yH6FB2AfA'
    gpais = ''
    gestado = ''
    gciudad = ''
    gcolonia = ''
    gcp = ''
    glat = ''
    glng = ''
    cont = 0
    data = []
    idZonaAct = 0

    geo = Geocercas.objects.filter(Q(estatus=1)).filter(Q(lng='')).filter(Q(lat=''))

    for g in geo:
        idZonaAct = g.idGeocerca
        direccion_codificada_origen = requests.utils.quote(g.direccion)

        response_origen = requests.get(f"https://maps.googleapis.com/maps/api/geocode/json?address={direccion_codificada_origen}&key={API_KEY}")
        # Verificar si la solicitud fue exitosa
        if response_origen.status_code == 200:
            data = response_origen.json()
            #print(data)
            # Verificar si se encontró una ubicación
            if data.get('results') and len(data['results']) > 0:
                glat = data['results'][0]['geometry']['location']['lat']
                glng = data['results'][0]['geometry']['location']['lng']

                act = Geocercas.objects.get(idGeocerca = idZonaAct)

                act.lng = glng
                act.lng_centro = glng
                act.lat = glat
                act.lat_centro = glat

                act.save()

    return Response({"info": ''})
    
